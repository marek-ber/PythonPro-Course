# 9. Prosty arkusz kalkulacyjny: Używając openpyxl , stwórz plik finanse.xlsx . W
# pierwszej kolumnie umieść nazwy wydatków (np. "Czynsz", "Jedzenie"), a w drugiej ich
# wartości. W komórce poniżej wartości oblicz i wstaw sumę wszystkich wydatków, używając
# formuły Excela (np. =SUM(B1:B2) ).

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

wydatki = [
    ("Czynsz", 1000),
    ("Jedzenie", 500)
    ]

ws["A1"] = "Wydatek"
ws["B1"] = "Kwota"

for i, (nazwa, kwota) in enumerate(wydatki, start=2):
    ws.cell(row=i, column=1, value=nazwa)
    ws.cell(row=i, column=2, value=kwota)


last_data_row = len(wydatki) + 1  
sum_row = last_data_row + 1       

ws[f"A{sum_row}"] = "Suma"
ws[f"B{sum_row}"] = f"=SUM(B2:B{last_data_row})"

wb.save("finanse.xlsx")